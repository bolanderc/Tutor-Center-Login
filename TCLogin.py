# -*- coding: utf-8 -*-
"""The login system for the Utah State University ENGR Tutor Center.

This login system uses Tkinter to create a user interface for collecting and
storing login information. This system is meant to be easily modified as the
needs of the Tutor Center and its management change.

Routine Listings
-----------------
LoginSystem     Class housing the login system GUI (graphical user interface).

Notes
------
This module is for running the Engineering Tutor Center Login System. Most of
the changes made to this system should be made in the CourseInfo.py module. The
idea of this documentation is to ensure that this system is easily modifiable
to the future needs of the Tutor Center. There are in-line comments throughout
the script that describe what each chunk of code does, so it should be easy to
navigate to a section that you would like to change and make those changes.

***IMPORTANT: Please verify all changes with the Tutor Center Supervisor before
they are made.

To run the login system, simply run this script!

"""

import CourseInfo  # CourseInfo.py must be in the same directory as this script
import datetime
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.messagebox as messagebox
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas


class LoginSystem:
    """LoginSystem is the class that houses the entire GUI.

    This class forms the entire GUI with its accompanying functionality. In
    general, there should be no need to change this section of code. Rather, as
    new classes are needed, they should be added to the CourseInfo class
    following the instructions contained therein. If the need arises to change
    this system, the following documentation should allow such changes to be
    easily made.

    The __init__ constructor will run each time an instance of the class is
    created. It sets up the GUI initially, and relies on the `name_change` and
    `record_data` methods for modifications to the original setup. These
    methods affect `name_menu` and the final data output respectively.

    Attributes
    ------------
    anumber_entry : tkinter.ttk.Entry
        Widget for taking a user entry for A number.
    anumberlabel : tkinter.ttk.Label
        Widget for holding the label for A-number.
    courselabel : tkinter.ttk.Label
        Widget for holding the label for course prefix.
    coursenamelabel : tkinter.ttk.Label
        Widget for holding the label for course name.
    courses
    friday_label : tkinter.ttk.Label
        Widget for holding a label explaining the Friday hours of the Tutor
        Center.
    majorlabel : tkinter.ttk.Label
        Widget for holding the label for major.
    major_menu : tkinter.OptionMenu
        An option menu containing all of the majors in `major_options`.
    major_options : CourseInfo
        Holds all of the options available to choose a major.
    major_selection : tkinter.StringVar
        Contains the selected major from `major_menu`.
    majorvar : tkinter.StringVar
        A variable that changes every time a new major is selected from
        `major_menu`.
    name_menu : tkinter.OptionMenu
        An option menu containing all of the course prefixes in `name_options`.
    name_options : CourseInfo
        Holds all of the options available to choose course name.
    name_selection : tkinter.StringVar
        Contains the selected course name from `name_menu`.
    namevar : tkinter.StringVar
        A variable that changes every time a new course name is selected from
        `name_menu`.
    prefix_menu : tkinter.OptionMenu
        An option menu containing all of the course prefixes in
        `prefix_options`.
    prefix_options : CourseInfo
        Holds all of the options available to choose course prefix.
    prefix_selection : tkinter.StringVar
        Contains the selected course prefix from `prefix_menu`.
    prefixvar : tkinter.StringVar
        A variable that changes every time a new course prefix is selected from
        `prefix_menu`.
    ranklabel : tkinter.ttk.Label
        Widget for holding the label for class rank.
    rank_menu : tkinter.OptionMenu
        An option menu containing all of the class ranks in `rank_options`.
    ranknote : string
        Text to be placed at the bottom of the page as a note for class rank.
    rank_options : CourseInfo
        Holds all of the options available to choose class rank.
    rank_selection : tkinter.StringVar
        Contains the selected class rank from `rank_menu`.
    rank_sublabel : tkinter.ttk.Label
        Widget for holding a label at the bottom of the page with a note about
        class rank.
    rankvar : tkinter.StringVar
        A variable that changes every time a new class rank is selected from
        `rank_menu`.
    record_button : tkinter.Button
        Widget that creates a button to record the data input by the student.
    side_bar_title : tkinter.ttk.Label
        Widget for holding a label for the Tutor Center hours in the sidebar.
    weekday_label : tkinter.ttk.Label
        Widget for holding a label explaining the Monday-Thursday hours of the
        Tutor Center.
    weekend_label : tkinter.ttk.Label
        Widget for holding a label explaining the Saturday-Sunday hours of the
        Tutor Center.

    See Also
    -----------
    CourseInfo.py : Module containing the class for storing course information.

    tkinter.Entry : Creates a widget for string entry.
    tkinter.Label : Creates a label widget.
    tkinter.OptionMenu : Creates a pre-configured option menu.
    tkinter.PhotoImage : Imports an image to be used in a widget.
    tkinter.StringVar : Creates a string variable to be used in a widget.

    tkinter.ttk : Allows for more creative changes to tkinter widgets.

    """

    def __init__(self, master):
        # Sets up the root window of the GUI with a title, the proper size,
        # a background color and a logo. Also defines an instance of the
        # CourseInfo class that will be used in the dropdown menus in the GUI.
        self.master = master
        master.title("ENGR Tutor Center Login")
        master.geometry("%dx%d+0+0" % (root.winfo_screenwidth(),
                        root.winfo_screenheight()))
        master.configure(background="silver")
        master.iconbitmap('Logo.ico')
        self.courses = CourseInfo.CourseInfo()

        # Sets up the welcome banner from an image contained in the folder
        # where the GUI is stored.
        self.welcome_image = tk.PhotoImage(file="Welcome.gif")
        self.welcomelabel = tk.Label(master, image=self.welcome_image)
        self.welcomelabel.config(borderwidth=0)

        # Saves a reference of the welcome_image class variable. This is very
        # important! If this reference is not kept, then everytime the GUI is
        # closed it will throw out the image reference and not be able to find
        # it if the GUI is opened again.
        self.welcomelabel.image = self.welcome_image

        # Defines all main data labels with their text, font, background
        # and letter colors. The foreground color is USU blue in hexidecimal.
        self.anumberlabel = ttk.Label(master,
                                      text='A-Number',
                                      font='Helvetica 20 bold',
                                      background='silver',
                                      foreground='#0F2439')
        self.anumber_entry = ttk.Entry(master, background="white")
        self.majorlabel = ttk.Label(master,
                                    text="Major",
                                    font='Helvetica 20 bold',
                                    background='silver',
                                    foreground='#0F2439')
        self.ranklabel = ttk.Label(master,
                                   text="Class Rank *",
                                   font='Helvetica 20 bold',
                                   background='silver',
                                   foreground='#0F2439')
        self.courselabel = ttk.Label(master,
                                     text="Course Prefix",
                                     font='Helvetica 20 bold',
                                     background='silver',
                                     foreground='#0F2439')
        self.coursenamelabel = ttk.Label(master,
                                         text="Course Name",
                                         font='Helvetica 20 bold',
                                         background='silver',
                                         foreground='#0F2439')
        self.ranknote = ("*Note that Class Rank is measured by the courses you"
                         " are taking, not the number of credits you have.")
        self.rank_sublabel = ttk.Label(master, text=self.ranknote,
                                       font='Helvetica 12 bold',
                                       background='silver',
                                       foreground='#0F2439',
                                       wraplength=500,
                                       justify=tk.CENTER)

        # Creates a button that will be used by the student to record their
        # information and sign in to the Tutor Center. The `record_data` method
        # will be run everytime this button is pressed.
        self.record_button = tk.Button(master, text="Sign In",
                                       font="Helvetica 16 bold",
                                       background="#0F2439",
                                       foreground="white",
                                       command=self.record_data)

        # Sets up the labels for the Tutor Center hours that will be shown in
        # the left sidebar with their text, font, background and foreground
        # colors.
        self.side_bar_title = ttk.Label(master,
                                        text="Tutor Center Hours",
                                        font='Helvetica 20 bold',
                                        background='silver',
                                        foreground='#0F2439')

        self.weekday_label = ttk.Label(master, text=("Monday - Thursday\n"
                                                     "8 AM - 7 PM"),
                                       font='Helvetica 14 bold',
                                       background='silver',
                                       foreground='#0F2439',
                                       justify=tk.CENTER)
        self.friday_label = ttk.Label(master, text="Friday\n8 AM - 4 PM",
                                      font='Helvetica 14 bold',
                                      background='silver',
                                      foreground='#0F2439',
                                      justify=tk.CENTER)
        self.weekend_label = ttk.Label(master, text="Closed Saturday & Sunday",
                                       font='Helvetica 12 bold',
                                       background='silver',
                                       foreground='#0F2439',
                                       justify=tk.CENTER)

        # Accesses the drop-down options stored in the CourseInfo class and
        # assigns them to a class variable. These will be the options stored
        # in the drop-down buttons created later.
        self.major_options = self.courses.majoroptions
        self.rank_options = self.courses.rankoptions
        self.prefix_options = self.courses.prefixoptions

        # Major Selection Menu:
        # The majorvar variable is used for tracking the student's selection
        # of their major based on the major menu.
        self.majorvar = tk.StringVar(root)
        self.majorvar.set(self.major_options[0])  # Sets a default major.

        # major_menu is the button that the student will interact with to
        # select their major. major_selection stores the value the student
        # chooses for their major.
        self.major_menu = tk.OptionMenu(master, self.majorvar,
                                        *self.major_options)
        self.major_selection = self.majorvar.get()
        self.major_menu.config(bg="white", font='Helvetica 12',
                               foreground="#0F2439",
                               highlightthickness=0)
        self.major_menu["menu"].config(bg="white")

        # Class Rank Selection Menu:
        # The rankvar variable is used for tracking the student's selection
        # of their class rank based on the class rank menu.
        self.rankvar = tk.StringVar(master)
        self.rankvar.set(self.rank_options[0])  # Sets a default class rank.

        # rank_menu is the button that the student will interact with to
        # select their class rank. rank_selection stores the value the student
        # chooses for their class rank.
        self.rank_menu = tk.OptionMenu(master, self.rankvar,
                                       *self.rank_options)
        self.rank_selection = self.rankvar.get()
        self.rank_menu.config(bg="white", font='Helvetica 12',
                              foreground="#0F2439",
                              highlightthickness=0)
        self.rank_menu["menu"].config(bg="white")

        # Course Prefix Selection Menu:
        # The prefixvar variable is used for tracking the student's selection
        # of their course prefix based on the course prefix menu.
        self.prefixvar = tk.StringVar(master)
        self.prefixvar.set(self.prefix_options[0])  # Sets a default.

        # prefix_menu is the button that the student will interact with to
        # select their course prefix. prefix_selection stores the value the
        # student chooses for their course prefix.
        self.prefix_menu = tk.OptionMenu(master, self.prefixvar,
                                         *self.prefix_options)
        self.prefix_selection = self.prefixvar.get()
        self.prefix_menu.config(bg="white", font='Helvetica 12',
                                foreground="#0F2439",
                                highlightthickness=0)
        self.prefix_menu["menu"].config(bg="white")

        # Used to populate name_menu with the correct options, based on the
        # course prefix. Links the chosen prefix to the list of possible
        # courses using that prefix.
        self.name_options = self.courses.populate_names(self.prefix_selection)

        # Course Name Selection Menu:
        # The namevar variable is used for tracking the student's selection
        # of their course name based on the course name menu.
        self.namevar = tk.StringVar(master)
        self.namevar.set(self.name_options[0])

        # name_menu is the button that the student will interact with to
        # select their course name. name_selection stores the value the
        # student chooses for their course name.
        self.name_menu = tk.OptionMenu(master, self.namevar,
                                       *self.name_options)
        self.name_menu.config(bg="white", font='Helvetica 12',
                              foreground="#0F2439",
                              highlightthickness=0)
        self.name_menu["menu"].config(bg="white")
        # A trace for prefixvar is set up here. This trace employs the method
        # `name_change` everytime a new value is written (selected by the
        # student, hence the "w" for write) to prefixvar.
        self.prefixvar.trace("w", self.name_change)
        self.name_selection = self.namevar.get()

        # Creates an 8x5 grid in which to place each of the labels, buttons,
        # etc.
        for r in range(7):
            master.rowconfigure(r, weight=1)
        for c in range(4):
            master.columnconfigure(c, weight=1)

        # Sets the position of each of the labels, menus, and buttons in the
        # GUI.
        self.welcomelabel.grid(row=0, column=0, columnspan=5)
        self.side_bar_title.grid(row=1, column=0)
        self.anumberlabel.grid(row=1, column=1)
        self.anumber_entry.grid(row=1, column=2)
        self.weekday_label.grid(row=2, column=0)
        self.majorlabel.grid(row=2, column=1)
        self.major_menu.grid(row=2, column=2)
        self.friday_label.grid(row=3, column=0)
        self.ranklabel.grid(row=3, column=1)
        self.rank_menu.grid(row=3, column=2)
        self.weekend_label.grid(row=4, column=0)
        self.courselabel.grid(row=4, column=1)
        self.prefix_menu.grid(row=4, column=2)
        self.coursenamelabel.grid(row=5, column=1)
        self.name_menu.grid(row=5, column=2)
        self.rank_sublabel.grid(row=6, column=1)
        self.record_button.grid(row=6, column=2)

    def name_change(self, *args):
        """Changes the list of options in `name_menu` based on user input.

        The selected prefix is retrieved from `prefixvar` and a new set of
        course names is populated and replaces the values in `name_menu`.

        Returns
        -------
        name_menu : tkinter.OptionMenu
            Option menu with the course names corresponding to a selected
            course prefix.

        See Also
        --------
        CourseInfo.py

        """

        self.prefix_chosen = self.prefixvar.get()
        self.new_names = self.courses.populate_names(self.prefix_chosen)
        self.namevar.set(self.new_names[0])
        # Deletes previous options in name_menu ...
        self.name_menu['menu'].delete(0, 'end')
        # ... and replaces them with the options in `new_names`.
        for names in self.new_names:
            self.name_menu['menu'].add_command(label=names,
                                               command=tk._setit(self.namevar,
                                                                 names))
        return self.name_menu

    def record_data(self):
        """Records data input by user into the GUI.

        This method records the information that is input by the student and
        saves it in the Masterfile.xlsx Excel spreadsheet. This spreadsheet
        must be located in the same directory as this script. Please refer to
        the documentation for the spreadsheet for questions of upkeep.

        See Also
        --------
        tkinter.messagebox : Creates a messagebox widget to display text after
                             an event

        datetime.datetime : Returns time and date information.

        openpyxl.load_workbook : Loads Excel workbook for use in saving data
        openpyxl.utils.datafram.dataframe_to_rows : Takes a dataframe and
                                                    places it in a row on a
                                                    spreadsheet.

        pandas.DataFrame : Creates a dataframe (like a dictionary) for storing
                           data.
        """

        # Checks if A-number input is correct by checking the string length and
        # that the A-number field starts wth "A" or "a".
        self.a_get = self.anumber_entry.get()
        self.a_length = len(self.anumber_entry.get())
        if self.a_length == 9 and (self.a_get.startswith('A')
                                   or self.a_get.startswith('a')):

            # Grabs information for the current date and time. This data is
            # then saved as a full date in the form of `Day of the Week`,
            # `Month` `Day`, `Year` in the full_date variable. The day of the
            # week is stored (Mon-Sun), as well as the time that the student
            # entered the Tutor Center (Hour:Minutes AM/PM).
            self.date_and_time = datetime.datetime.now()
            self.full_date = self.date_and_time.strftime("%A,%B %d,%Y")
            self.day_of_week = self.date_and_time.strftime("%A")
            self.timein = self.date_and_time.strftime("%I:%M %p")

            # Loads the Masterfile spreadsheet and selects the active sheet to
            # record the student's data.
            self.wb = load_workbook('Masterfile.xlsx')
            self.ws = self.wb["Main Data"]

            # Gets all of the data given by the student, saves it as a pandas
            # dataframe (dictionary_like), and then adds the values to the
            # current worksheet in Masterfile.xlsx.
            self.a_number = self.anumber_entry.get()
            self.major = self.majorvar.get()
            self.class_rank = self.rankvar.get()
            self.course_prefix = self.prefixvar.get()
            self.course_name = self.namevar.get()
            self.data = pandas.DataFrame({'Date': [self.full_date],
                                          'Day': [self.day_of_week],
                                          'Time In': [self.timein],
                                          'Anumber': [self.a_number],
                                          'Major': [self.major],
                                          'Class Rank': [self.class_rank],
                                          'Course Prefix':
                                              [self.course_prefix],
                                          'Course Name': [self.course_name],
                                          }, columns=['Anumber', 'Class Rank',
                                                      'Major', 'Course Prefix',
                                                      'Course Name', 'Date',
                                                      'Day', 'Time In'])
            # Appends data to Masterfile.
            for r in dataframe_to_rows(self.data,
                                       index=False, header=False):
                self.ws.append(r)

            # Re-saves the Excel file full of data and clears the A-number from
            # anumber_entry.
            self.wb.save('Masterfile.xlsx')
            self.anumber_entry.delete(0, 'end')
            messagebox.showinfo("Login Confirmation", "Thank you!")

        # If the A-Number entered does not start with "A"/"a" or is not long
        # enough, prompts the student to change their A-Number input.
        else:
            messagebox.showinfo("A-Number Error",
                                "Please check your A-Number and try again.")


# This section executes the GUI. It creates a root window for the application
# to be run in, and then places all of the widgets and functionality defined
# in the LoginSystem class in that root window. It then loops that root window
# continuously until the user exits the window.
root = tk.Tk()
Login = LoginSystem(root)
root.mainloop()
